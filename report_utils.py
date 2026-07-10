"""
Styled Excel writer for the climate-indices summary table.
Pure representation/formatting layer -- the values themselves come from
main.py's `rows` list exactly as computed; nothing here changes any number,
only how it's displayed (colors, units, number formats, layout).
"""
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT = '1B6E64'          # matches the web app's teal accent
HEADER_FILL = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color=ACCENT)
SUBTITLE_FONT = Font(name='Arial', italic=True, size=9, color='6B6B6B')
LEGEND_FONT = Font(name='Arial', size=8, color='6B6B6B')
BASE_FONT = Font(name='Arial', size=10)
HEADLINE_FONT = Font(name='Arial', size=10, bold=True)

BASELINE_FILL = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
TEMP_FILL = PatternFill(start_color='FDECEA', end_color='FDECEA', fill_type='solid')     # warm tint
PRECIP_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')   # cool tint

_thin = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

NUMERIC_COLS = ('mean', 'p10', 'p90', 'headline_value')
COLUMN_ORDER = ['period', 'domain', 'index', 'unit', 'mean', 'p10', 'p90', 'headline_value']


def _infer_unit(domain, idx_name):
    name = idx_name.lower()
    if name.startswith('gev_return_levels'):
        return 'mm'
    if domain == 'temperature':
        return 'days' if 'su_days_per_month' in name else '°C'
    if domain == 'precipitation':
        return 'days' if name.startswith('wetdays_per_month') else 'mm'
    return ''


def write_excel_summary(df, out_path, meta):
    """
    df: DataFrame with columns ['period','domain','index','mean','p10','p90','headline_value'].
    meta: dict with optional keys 'baseline_period', 'models', 'scenarios', 'generated'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Climate Indices Summary'

    n_cols = len(COLUMN_ORDER)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'Climate Indices Summary'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    subtitle = (
        f"Baseline {meta.get('baseline_period', '')}  ·  "
        f"Models: {', '.join(meta.get('models', []))}  ·  "
        f"Scenarios: {', '.join(s.upper() for s in meta.get('scenarios', []))}  ·  "
        f"Generated {meta.get('generated', datetime.now().strftime('%Y-%m-%d %H:%M'))}"
    )
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = subtitle
    ws['A2'].font = SUBTITLE_FONT

    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'] = ("Row shading — gray: baseline (no ensemble spread) · red tint: temperature "
                "index · blue tint: precipitation index. 'Headline Value' is the statistic "
                "(mean/p10/p90) configured as each index's representative value.")
    ws['A3'].font = LEGEND_FONT
    ws['A3'].alignment = Alignment(wrap_text=True)

    header_row = 5
    for c, col_name in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=header_row, column=c, value=col_name.replace('_', ' ').title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    col_max_len = {col: len(col.replace('_', ' ').title()) for col in COLUMN_ORDER}

    for r_offset, row in enumerate(df.itertuples(index=False), start=1):
        row_dict = row._asdict()
        domain = row_dict.get('domain', '')
        period = row_dict.get('period', '')
        idx_name = row_dict.get('index', '')
        row_dict['unit'] = _infer_unit(domain, idx_name)

        r = header_row + r_offset
        if period == 'Baseline':
            row_fill = BASELINE_FILL
        elif domain == 'temperature':
            row_fill = TEMP_FILL
        elif domain == 'precipitation':
            row_fill = PRECIP_FILL
        else:
            row_fill = None

        for c, col_name in enumerate(COLUMN_ORDER, start=1):
            val = row_dict.get(col_name)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = HEADLINE_FONT if col_name == 'headline_value' else BASE_FONT
            cell.border = THIN_BORDER
            if row_fill is not None:
                cell.fill = row_fill
            if col_name in NUMERIC_COLS:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            col_max_len[col_name] = max(col_max_len[col_name], len(str(val)) if val is not None else 0)

    for c, col_name in enumerate(COLUMN_ORDER, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(col_max_len[col_name] + 3, 10), 42)

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{header_row}'
    ws.sheet_view.showGridLines = False

    wb.save(out_path)
