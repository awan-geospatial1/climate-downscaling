"""
template_excel_utils.py - Builds a workbook matching the layout of the
provided Template.xlsx: Daily Spatial Averages, Helper, Precipitation
Stats and Graph, Max_Precip, Return Period Graphs, Temperature Stats and
Graphs.

IMPORTANT - please read before treating this as a finished, pixel-perfect
match:

The actual Template.xlsx has 13 native Excel chart objects across 3 of
its 6 sheets (Precipitation Stats and Graph: 3, Return Period Graphs: 8,
Temperature Stats and Graphs: 2), plus header/merge structure deeper than
what's mechanically derivable from the pipeline's existing outputs alone.
This module reproduces the DATA layer of all 6 sheets faithfully (same
sheet names, same column/row groupings, same source-of-truth values) and
adds one straightforward chart per data sheet where a chart makes sense,
but it is a first pass on layout/chart styling, not a guaranteed
cell-for-cell match to the original template's exact formatting or every
one of its 13 charts. Worth reviewing against the template before treating
this as final -- flagged explicitly rather than silently pretending it's
an exact clone.

Two specific judgment calls, documented here rather than buried in a
comment nobody will find:

1. Max_Precip / Return Period Graphs use an "ensemble max" that takes
   each model's own AOI spatial-mean daily precip series FIRST, then the
   maximum across models for each day -- NOT a spatial average of the
   cellwise ensemble-max grid saved by ensemble_utils.compute_ensemble_max
   (max and spatial averaging don't commute, so these differ). This
   series-level version is what an AOI-level daily time series /
   extreme-value table should be built from; the cellwise grid remains
   the right choice for spatial hazard maps.

2. Return Period Graphs reports Gumbel method-of-moments parameters
   (n, xbar, st.dev, alpha, u), matching the exact statistic labels shown
   in the template. This is a DIFFERENT, simpler method from the
   GEV-with-bootstrap approach already used elsewhere in this pipeline
   (indices_utils.gev_return_levels, feeding climate_indices_summary.xlsx)
   -- both are legitimate, standard extreme-value approaches, but they are
   not the same method and will not produce identical numbers. This
   module does not touch or replace the existing GEV output.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from indices_utils import daily_spatial_series, daily_spatial_ensemble

HEADER_FILL = PatternFill('solid', fgColor='118D9E')
HEADER_FONT = Font(color='FFFFFF', bold=True)
SUBHEADER_FILL = PatternFill('solid', fgColor='118D9E')
CENTER = Alignment(horizontal='center', vertical='center')
THIN_SIDE = Side(style='thin', color='BFBFBF')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


# =========================================================================
# Small shared helpers
# =========================================================================
def daily_spatial_ensemble_max(grids_by_model, units=None):
    """Mirrors indices_utils.daily_spatial_ensemble, but takes the MAXIMUM
    across models' own spatial-mean daily series instead of the mean --
    see judgment call #1 in the module docstring above."""
    if not grids_by_model:
        return None
    series_list = [daily_spatial_series(da, units) for da in grids_by_model.values()]
    return pd.concat(series_list, axis=1).max(axis=1)


def _concat_periods(series_by_tag):
    """dict[tag] -> pd.Series (or None), sorted by their own date index,
    concatenated into one continuous series. Drops any None entries
    (periods where every model failed)."""
    parts = [s for s in series_by_tag.values() if s is not None]
    if not parts:
        return None
    return pd.concat(parts).sort_index()


def _gumbel_moments(annual_max_values):
    """Classic 2-parameter Gumbel, method of moments:
      alpha (scale) = sqrt(6) * s / pi
      u     (location) = xbar - 0.5772156649 (Euler-Mascheroni) * alpha
    This is the method the template's 'n / xbar / st.dev / alpha / u'
    row labels correspond to -- see judgment call #2 above for how this
    relates to (and differs from) the pipeline's existing GEV output."""
    x = np.asarray(annual_max_values, dtype=float)
    x = x[np.isfinite(x)]
    n = len(x)
    if n < 2:
        return dict(n=n, xbar=np.nan, stdev=np.nan, alpha=np.nan, u=np.nan)
    xbar = float(np.mean(x))
    stdev = float(np.std(x, ddof=1))
    alpha = float(np.sqrt(6) * stdev / np.pi)
    u = float(xbar - 0.5772156649 * alpha)
    return dict(n=n, xbar=xbar, stdev=stdev, alpha=alpha, u=u)


def _style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _apply_borders(ws, min_row, max_row, min_col, max_col):
    """Thin grey borders on every cell in the range — matches the
    reference report's tables, which border every cell including the
    plain data rows (not just headers). Added across the main summary
    tables below; previously no border styling existed anywhere in this
    file, which read as noticeably plainer/less finished than the
    reference report's Excel-native table look."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def _period_labels(future_intervals):
    """[(tag, 'Short (2021-2040)'), ...] -- combines each interval's label
    and tag into the display string the template's headers use."""
    return [(tag, f'{label} ({tag})') for (_, _, label, tag) in future_intervals]


# =========================================================================
# Sheet 1: Daily Spatial Averages
# =========================================================================
def _sheet_daily_spatial_averages(wb, variables, _cfg, scenarios, future_intervals,
                                   ref_cache, corrected_grids):
    ws = wb.create_sheet('Daily Spatial Averages')

    groups = {'Baseline': None}
    for scenario in scenarios:
        groups[scenario.upper()] = scenario

    frames = {}
    base_cols = {}
    for var in variables:
        try:
            base_cols[var] = daily_spatial_series(ref_cache[var]['ref'], _cfg[var]['ref_units'])
        except Exception as e:
            print(f"    [template-excel] baseline series for {var} failed: {e}")
    frames['Baseline'] = pd.DataFrame(base_cols) if base_cols else pd.DataFrame()

    for scenario in scenarios:
        cols = {}
        for var in variables:
            try:
                by_tag = {tag: daily_spatial_ensemble(corrected_grids[var][scenario][tag], _cfg[var]['ref_units'])
                          for (_, _, _, tag) in future_intervals}
                s = _concat_periods(by_tag)
                if s is not None:
                    cols[var] = s
            except Exception as e:
                print(f"    [template-excel] {scenario} series for {var} failed: {e}")
        frames[scenario.upper()] = pd.DataFrame(cols) if cols else pd.DataFrame()

    # Header row 1: group names (merged across each group's variable columns)
    # Header row 2: variable names. Column A: Date.
    ws.cell(row=1, column=1, value='Date')
    ws.cell(row=2, column=1, value='')
    col = 2
    for group_name, frame in frames.items():
        if frame.empty:
            continue
        start_col = col
        for var in variables:
            if var not in frame.columns:
                continue
            ws.cell(row=2, column=col, value=var)
            col += 1
        end_col = col - 1
        if end_col >= start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col, value=group_name)
    _style_header_row(ws, 1, 1, col - 1)
    _style_header_row(ws, 2, 1, col - 1)

    # Union of all dates across groups, sorted, one row per date.
    all_dates = sorted(set().union(*[frame.index for frame in frames.values() if not frame.empty]))
    row = 3
    for d in all_dates:
        ws.cell(row=row, column=1, value=pd.Timestamp(d).date())
        col = 2
        for group_name, frame in frames.items():
            if frame.empty:
                continue
            for var in variables:
                if var not in frame.columns:
                    continue
                val = frame[var].get(d, None)
                if val is not None and pd.notna(val):
                    ws.cell(row=row, column=col, value=round(float(val), 3))
                col += 1
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.freeze_panes = 'B3'


# =========================================================================
# Sheet 2: Temperature Stats and Graphs
# =========================================================================
def _add_monthly_seasonality_block(ws, results, scenarios, periods, category_key, var_key,
                                    row_start, y_axis_title, chart_title, chart_anchor_col):
    """
    Monthly climatology table + smoothed line chart: Baseline plus one line
    per scenario/period, matching the reference report's monthly
    seasonality charts (temperature and precipitation both have one).

    FIX: monthly_mean_tas already existed as a computed index, but nothing
    in this file ever laid it out as a table or charted it — and
    monthly_mean_pr didn't exist at all until this same pass (see
    indices_utils.py). Both variables get identical treatment here via
    category_key ('temperature'/'precipitation') and var_key
    ('monthly_mean_tas'/'monthly_mean_pr').
    """
    ws.cell(row=row_start, column=1, value='Month')
    for j, m in enumerate(MONTH_NAMES):
        ws.cell(row=row_start, column=2 + j, value=m)
    _style_header_row(ws, row_start, 1, 1 + len(MONTH_NAMES))

    row = row_start + 1
    table_top = row

    def write_monthly_row(label, monthly_list):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        if monthly_list:
            for j, v in enumerate(monthly_list[:12]):
                if v is not None:
                    ws.cell(row=row, column=2 + j, value=round(float(v), 2))
        row += 1

    base_stat = results.get('Baseline', {}).get(category_key, {}).get(var_key, {})
    base_vals = base_stat.get('mean') if isinstance(base_stat, dict) else base_stat
    write_monthly_row('Baseline', base_vals)

    for scenario in scenarios:
        for tag, label in periods:
            stat = results.get(f'{scenario}_{tag}', {}).get(category_key, {}).get(var_key, {})
            vals = stat.get('mean') if isinstance(stat, dict) else None
            write_monthly_row(f'{scenario.upper()} {label}', vals)

    table_bottom = row - 1
    _apply_borders(ws, row_start, table_bottom, 1, 1 + len(MONTH_NAMES))

    try:
        chart = LineChart()
        chart.title = chart_title
        chart.y_axis.title = y_axis_title
        chart.x_axis.title = 'Month'
        chart.style = 2
        # FIX: this used to start at table_top - 1 (the "Month/Jan/Feb/..."
        # header row itself), which got misread as its own data series
        # ("Month", flat at 0, since the month names aren't numbers) — a
        # phantom line visible in the rendered chart. from_rows=True with
        # titles_from_data pulls each row's own column-A value as that
        # row's series title directly, so the header row must NOT be
        # included in the data range at all -- start at table_top (the
        # first real data row, "Baseline") instead.
        data = Reference(ws, min_col=1, max_col=1 + len(MONTH_NAMES), min_row=table_top, max_row=table_bottom)
        # from_rows=True: each ROW (Baseline, each scenario/period) is one
        # series, months run across columns -- opposite orientation from
        # every other chart in this file, which is why this needs its own
        # Reference/add_data call instead of reusing the column-series helper
        # pattern the totals charts above use.
        chart.add_data(data, titles_from_data=True, from_rows=True)
        cats = Reference(ws, min_col=2, max_col=1 + len(MONTH_NAMES), min_row=row_start, max_row=row_start)
        chart.set_categories(cats)
        for s in chart.series:
            s.smooth = True
        ws.add_chart(chart, f'{get_column_letter(chart_anchor_col)}{row_start}')
    except Exception as e:
        print(f"    [template-excel] monthly seasonality chart failed ({chart_title}): {e}")

    return row + 2  # next free row, with a blank-row gap


def _sheet_temperature_stats(wb, results, scenarios, future_intervals):
    ws = wb.create_sheet('Temperature Stats and Graphs')
    periods = _period_labels(future_intervals)
    temp_vars = ['tas', 'tasmax', 'tasmin']

    # Header: A = metric, B = Baseline, then one column per scenario x period
    ws.cell(row=1, column=1, value='Spatial Average (\u00b0C)')
    ws.cell(row=1, column=2, value='Baseline')
    col = 3
    scen_start_col = {}
    for scenario in scenarios:
        scen_start_col[scenario] = col
        for _, label in periods:
            ws.cell(row=2, column=col, value=label)
            col += 1
        end_col = col - 1
        ws.merge_cells(start_row=1, start_column=scen_start_col[scenario], end_row=1, end_column=end_col)
        ws.cell(row=1, column=scen_start_col[scenario], value=scenario.upper())
    total_cols = col - 1
    _style_header_row(ws, 1, 1, total_cols)
    _style_header_row(ws, 2, 2, total_cols)

    row = 3
    chart_anchor_row = row
    for var in temp_vars:
        ws.cell(row=row, column=1, value=f'{var} (annual mean)')
        base_stat = results.get('Baseline', {}).get('temperature', {}).get(f'annual_mean_{var}', {})
        # NOTE: main.py wraps every SCALAR baseline index (not just scenario
        # indices) as {'mean': v, 'p10': v, 'p90': v} -- only LIST-valued
        # indices (e.g. monthly_mean_tas) stay raw. annual_mean_* is scalar,
        # so it's wrapped for Baseline too, same shape as scenario entries.
        base_val = base_stat.get('mean') if isinstance(base_stat, dict) else base_stat
        ws.cell(row=row, column=2, value=round(base_val, 2) if base_val is not None else None)
        col = 3
        for scenario in scenarios:
            for tag, _ in periods:
                key = f'{scenario}_{tag}'
                stat = results.get(key, {}).get('temperature', {}).get(f'annual_mean_{var}', {})
                val = stat.get('mean') if isinstance(stat, dict) else None
                ws.cell(row=row, column=col, value=round(val, 2) if val is not None else None)
                col += 1
        row += 1

    _apply_borders(ws, 1, row - 1, 1, total_cols)
    ws.column_dimensions['A'].width = 28

    # One simple bar chart comparing the three temperature variables'
    # annual means across all Baseline/scenario/period columns.
    try:
        chart = BarChart()
        chart.title = 'Spatial-average annual mean temperature'
        chart.y_axis.title = '\u00b0C'
        data = Reference(ws, min_col=2, max_col=total_cols, min_row=chart_anchor_row - 1,
                          max_row=chart_anchor_row + len(temp_vars) - 1)
        cats = Reference(ws, min_col=1, max_col=1, min_row=chart_anchor_row, max_row=chart_anchor_row + len(temp_vars) - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f'{get_column_letter(total_cols + 2)}2')
    except Exception as e:
        print(f"    [template-excel] temperature chart failed: {e}")

    _add_monthly_seasonality_block(
        ws, results, scenarios, periods, 'temperature', 'monthly_mean_tas',
        row_start=row + 2, y_axis_title='\u00b0C',
        chart_title='Monthly Mean Temperature — Baseline vs. Scenarios', chart_anchor_col=total_cols + 2)


# =========================================================================
# Sheet 3: Precipitation Stats and Graph
# =========================================================================
def _pick_threshold_key(precip_thresholds, target=20.0):
    if not precip_thresholds:
        return None, None
    closest = min(precip_thresholds, key=lambda t: abs(t - target))
    return closest, f'wetdays_per_month_{closest:g}mm'


def _sheet_precipitation_stats(wb, results, scenarios, future_intervals, precip_thresholds, return_periods=None):
    ws = wb.create_sheet('Precipitation Stats and Graph')
    periods = _period_labels(future_intervals)

    # ── Totals & extremes summary block ─────────────────────────────────
    # This block used to be entirely missing: compute_precipitation_indices()
    # / aggregate_across_models() already produce prcptot, wet_season_total,
    # dry_season_total and rx1day mean/p90 for every period, but this sheet
    # only ever wrote the per-threshold "wetdays/month" blocks below, so the
    # workbook never actually reported annual/seasonal totals or the 1-day
    # extreme amount anywhere — the Temperature sheet has its summary block,
    # Precipitation didn't. Mirrors that same layout.
    #
    # FIX: added rx5day (was computed nowhere at all until this same pass —
    # see indices_utils.py) and the GEV N-year return level, both present
    # in the reference report's equivalent table but missing here.
    totals_specs = [
        ('prcptot', 'Total annual precipitation (mm)'),
        ('wet_season_total', 'Wet-season total (mm)'),
        ('dry_season_total', 'Dry-season total (mm)'),
        ('rx1day_mean', 'Max 1-day precipitation, mean (mm)'),
        ('rx1day_p90', 'Max 1-day precipitation, p90 (mm)'),
        ('rx5day_mean', 'Max 5-day precipitation, mean (mm)'),
        ('rx5day_p90', 'Max 5-day precipitation, p90 (mm)'),
    ]
    ws.cell(row=1, column=1, value='Spatial Average')
    ws.cell(row=1, column=2, value='Baseline')
    col = 3
    scen_start_col = {}
    for scenario in scenarios:
        scen_start_col[scenario] = col
        for _, label in periods:
            ws.cell(row=2, column=col, value=label)
            col += 1
        end_col = col - 1
        ws.merge_cells(start_row=1, start_column=scen_start_col[scenario], end_row=1, end_column=end_col)
        ws.cell(row=1, column=scen_start_col[scenario], value=scenario.upper())
    total_cols = col - 1
    _style_header_row(ws, 1, 1, total_cols)
    _style_header_row(ws, 2, 2, total_cols)

    row = 3
    totals_anchor_row = row
    for key, row_label in totals_specs:
        ws.cell(row=row, column=1, value=row_label)
        base_stat = results.get('Baseline', {}).get('precipitation', {}).get(key, {})
        base_val = base_stat.get('mean') if isinstance(base_stat, dict) else base_stat
        ws.cell(row=row, column=2, value=round(base_val, 2) if base_val is not None else None)
        col = 3
        for scenario in scenarios:
            for tag, _ in periods:
                stat = results.get(f'{scenario}_{tag}', {}).get('precipitation', {}).get(key, {})
                val = stat.get('mean') if isinstance(stat, dict) else None
                ws.cell(row=row, column=col, value=round(val, 2) if val is not None else None)
                col += 1
        row += 1

    n_rows_in_table = len(totals_specs)
    gev_T = None
    if return_periods:
        gev_T = 100 if 100 in return_periods else max(return_periods)
        ws.cell(row=row, column=1, value=f'{gev_T}-year precipitation in 24-hour (mm)')
        base_gev = results.get('Baseline', {}).get('precipitation', {}).get('gev_return_levels', {}).get(gev_T, {})
        base_val = base_gev.get('mean') if isinstance(base_gev, dict) else None
        ws.cell(row=row, column=2, value=round(base_val, 2) if base_val is not None else None)
        col = 3
        for scenario in scenarios:
            for tag, _ in periods:
                gev = results.get(f'{scenario}_{tag}', {}).get('precipitation', {}).get('gev_return_levels', {}).get(gev_T, {})
                val = gev.get('mean') if isinstance(gev, dict) else None
                ws.cell(row=row, column=col, value=round(val, 2) if val is not None else None)
                col += 1
        row += 1
        n_rows_in_table += 1

    _apply_borders(ws, 1, totals_anchor_row + n_rows_in_table - 1, 1, total_cols)

    try:
        chart = BarChart()
        chart.title = 'Spatial-average precipitation totals & extremes'
        chart.y_axis.title = 'mm'
        data = Reference(ws, min_col=2, max_col=total_cols, min_row=totals_anchor_row - 1,
                          max_row=totals_anchor_row + n_rows_in_table - 1)
        cats = Reference(ws, min_col=1, max_col=1, min_row=totals_anchor_row,
                          max_row=totals_anchor_row + n_rows_in_table - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f'{get_column_letter(total_cols + 2)}2')
    except Exception as e:
        print(f"    [template-excel] precipitation totals chart failed: {e}")

    row_cursor = _add_monthly_seasonality_block(
        ws, results, scenarios, periods, 'precipitation', 'monthly_mean_pr',
        row_start=row + 2, y_axis_title='mm',
        chart_title='Monthly Mean Precipitation — Baseline vs. Scenarios', chart_anchor_col=total_cols + 2)

    if not precip_thresholds:
        ws.cell(row=row_cursor, column=1,
                value='No precip_thresholds configured - no wetdays-per-month blocks to report.')
        ws.column_dimensions['A'].width = 30
        return

    # Every configured threshold gets its own stacked block below, not just
    # whichever one happens to be closest to 20mm — the ensemble stats for
    # every threshold are already computed by aggregate_across_models, this
    # sheet was just silently dropping all but one of them.
    for thr in precip_thresholds:
        key = f'wetdays_per_month_{thr:g}mm'
        ws.cell(row=row_cursor, column=1, value=f'Days/month > {thr:g}mm')
        ws.cell(row=row_cursor, column=2, value='Baseline')
        col = 3
        header_row = row_cursor
        subheader_row = row_cursor + 1
        for scenario in scenarios:
            start_col = col
            for _, label in periods:
                ws.cell(row=subheader_row, column=col, value=label)
                col += 1
            ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=col - 1)
            ws.cell(row=header_row, column=start_col, value=scenario.upper())
        total_cols = col - 1
        _style_header_row(ws, header_row, 1, total_cols)
        _style_header_row(ws, subheader_row, 2, total_cols)

        base_stat = results.get('Baseline', {}).get('precipitation', {}).get(key, {})
        base_vals = base_stat.get('mean') if isinstance(base_stat, dict) else base_stat
        chart_start_row = subheader_row + 1
        for m in range(12):
            row = chart_start_row + m
            ws.cell(row=row, column=1, value=MONTH_NAMES[m])
            bv = base_vals[m] if isinstance(base_vals, list) else None
            ws.cell(row=row, column=2, value=round(bv, 2) if bv is not None else None)
            col = 3
            for scenario in scenarios:
                for tag, _ in periods:
                    stat = results.get(f'{scenario}_{tag}', {}).get('precipitation', {}).get(key, {})
                    vals = stat.get('mean') if isinstance(stat, dict) else None
                    v = vals[m] if isinstance(vals, list) else None
                    ws.cell(row=row, column=col, value=round(v, 2) if v is not None else None)
                    col += 1

        try:
            chart = LineChart()
            chart.title = f'Days/month with precip > {thr:g}mm'
            chart.y_axis.title = 'days'
            data = Reference(ws, min_col=2, max_col=total_cols, min_row=subheader_row, max_row=chart_start_row + 11)
            cats = Reference(ws, min_col=1, max_col=1, min_row=chart_start_row, max_row=chart_start_row + 11)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f'{get_column_letter(total_cols + 2)}{header_row}')
        except Exception as e:
            print(f"    [template-excel] precipitation chart ({thr:g}mm) failed: {e}")

        _apply_borders(ws, header_row, chart_start_row + 11, 1, total_cols)
        row_cursor = chart_start_row + 12 + 3  # blank rows before next threshold's block

    ws.column_dimensions['A'].width = 30


# =========================================================================
# Sheet 4: Max_Precip
# =========================================================================
def _sheet_max_precip(wb, scenarios, future_intervals, ref_cache, corrected_grids):
    ws = wb.create_sheet('Max_Precip')

    frames = {'Baseline': daily_spatial_series(ref_cache['pr']['ref'], 'mm/d').to_frame('pr')}
    for scenario in scenarios:
        by_tag = {tag: daily_spatial_ensemble_max(corrected_grids['pr'][scenario][tag], 'mm/d')
                  for (_, _, _, tag) in future_intervals}
        s = _concat_periods(by_tag)
        frames[scenario.upper()] = s.to_frame('pr') if s is not None else pd.DataFrame()

    ws.cell(row=1, column=1, value='Date')
    col = 2
    group_cols = {}
    for group_name, frame in frames.items():
        if frame.empty:
            continue
        ws.cell(row=1, column=col, value=group_name)
        group_cols[group_name] = col
        col += 1
    _style_header_row(ws, 1, 1, col - 1)

    all_dates = sorted(set().union(*[f.index for f in frames.values() if not f.empty]))
    row = 2
    for d in all_dates:
        ws.cell(row=row, column=1, value=pd.Timestamp(d).date())
        for group_name, c in group_cols.items():
            val = frames[group_name]['pr'].get(d, None)
            if val is not None and pd.notna(val):
                ws.cell(row=row, column=c, value=round(float(val), 2))
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.freeze_panes = 'B2'


# =========================================================================
# Sheet 5: Helper (year x month precip sums, one block per scenario)
# =========================================================================
def _sheet_helper(wb, scenarios, future_intervals, ref_cache, _cfg, corrected_grids):
    ws = wb.create_sheet('Helper')

    def year_month_sums(series):
        if series is None or series.empty:
            return pd.DataFrame()
        df = series.to_frame('pr')
        df['year'] = df.index.year
        df['month'] = df.index.month
        pivot = df.pivot_table(index='year', columns='month', values='pr', aggfunc='sum')
        pivot = pivot.reindex(columns=range(1, 13))
        pivot.columns = MONTH_NAMES
        return pivot

    blocks = {'Baseline': daily_spatial_series(ref_cache['pr']['ref'], 'mm/d')}
    for scenario in scenarios:
        by_tag = {tag: daily_spatial_ensemble(corrected_grids['pr'][scenario][tag], 'mm/d')
                  for (_, _, _, tag) in future_intervals}
        blocks[scenario.upper()] = _concat_periods(by_tag)

    row = 1
    for group_name, series in blocks.items():
        pivot = year_month_sums(series)
        ws.cell(row=row, column=1, value=f'{group_name} - monthly precip total (mm)')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        if pivot.empty:
            row += 2
            continue
        ws.cell(row=row, column=1, value='Year')
        for j, m in enumerate(MONTH_NAMES):
            ws.cell(row=row, column=2 + j, value=m)
        _style_header_row(ws, row, 1, 13)
        row += 1
        for year, vals in pivot.iterrows():
            ws.cell(row=row, column=1, value=int(year))
            for j, m in enumerate(MONTH_NAMES):
                v = vals[m]
                if pd.notna(v):
                    ws.cell(row=row, column=2 + j, value=round(float(v), 1))
            row += 1
        row += 2  # blank rows between blocks

    ws.column_dimensions['A'].width = 14


def _gumbel_return_level(u, alpha, T):
    """x(T) = u - alpha * ln(-ln(1 - 1/T)) -- standard Gumbel EVI quantile
    formula, using the same method-of-moments (u, alpha) already computed
    by _gumbel_moments above, so this curve is consistent with the stat
    block already in this sheet rather than a second, different estimate."""
    if not np.isfinite(u) or not np.isfinite(alpha) or alpha <= 0:
        return np.nan
    return u - alpha * np.log(-np.log(1 - 1.0 / T))


# =========================================================================
# Sheet 6: Return Period Graphs (Gumbel method-of-moments)
# =========================================================================
def _sheet_return_period(wb, scenarios, future_intervals, ref_cache, corrected_grids):
    ws = wb.create_sheet('Return Period Graphs')

    def annual_max(series, start=None, end=None):
        if series is None or series.empty:
            return np.array([])
        if start is not None:
            series = series[(series.index >= pd.Timestamp(start)) & (series.index <= pd.Timestamp(end))]
        return series.groupby(series.index.year).max().values

    blocks = []  # (label, annual_max_values)
    base_series = daily_spatial_series(ref_cache['pr']['ref'], 'mm/d')
    blocks.append(('Baseline', annual_max(base_series)))

    for scenario in scenarios:
        for start, end, label, tag in future_intervals:
            s = daily_spatial_ensemble_max(corrected_grids['pr'][scenario][tag], 'mm/d')
            blocks.append((f'{scenario.upper()} {label} ({tag})', annual_max(s, start, end)))

    STAT_ROWS = ['n', 'xbar', 'st. dev', 'alpha', 'u']
    col = 1
    gumbel_params = []  # (label, u, alpha) -- reused below for the curve table
    for label, vals in blocks:
        stats = _gumbel_moments(vals)
        gumbel_params.append((label, stats['u'], stats['alpha']))
        ws.cell(row=1, column=col, value=label)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = SUBHEADER_FILL
        for i, key in enumerate(('n', 'xbar', 'stdev', 'alpha', 'u')):
            ws.cell(row=2 + i, column=col, value=STAT_ROWS[i])
            ws.cell(row=2 + i, column=col + 1, value=round(stats[key], 3) if pd.notna(stats[key]) else None)
        # Annual-max series itself, below the stat block, for graphing.
        ws.cell(row=8, column=col, value='Annual max (mm/d)')
        for r, v in enumerate(vals):
            ws.cell(row=9 + r, column=col, value=round(float(v), 2))
        col += 3  # 2 cols for this block + 1 blank spacer

    # ── Return-level curve table + chart ────────────────────────────────
    # FIX: this sheet only ever stored the raw stats and annual-max series
    # needed to BUILD a return-period curve by hand in Excel -- it never
    # actually produced the curve or a chart, unlike the reference report's
    # "Precipitation vs Return Period" line chart (smooth curve, 2-100
    # years, one line per baseline/scenario-period). Fixed row 150 anchor:
    # safely below any realistic annual-max series length (even a full
    # 1990-2100 span is ~110 rows) so this never collides with the blocks
    # above regardless of how many years of data they contain.
    curve_row0 = 150
    return_periods_curve = [2, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    ws.cell(row=curve_row0, column=1, value='Return Period (years)')
    ws.cell(row=curve_row0, column=1).font = Font(bold=True)
    for j, (label, u, alpha) in enumerate(gumbel_params):
        ws.cell(row=curve_row0, column=2 + j, value=label)
    _style_header_row(ws, curve_row0, 1, 1 + len(gumbel_params))

    for i, T in enumerate(return_periods_curve):
        r = curve_row0 + 1 + i
        ws.cell(row=r, column=1, value=T)
        for j, (label, u, alpha) in enumerate(gumbel_params):
            val = _gumbel_return_level(u, alpha, T)
            ws.cell(row=r, column=2 + j, value=round(val, 2) if np.isfinite(val) else None)

    _apply_borders(ws, curve_row0, curve_row0 + len(return_periods_curve), 1, 1 + len(gumbel_params))

    try:
        chart = LineChart()
        chart.title = 'Precipitation vs. Return Period'
        chart.y_axis.title = 'Precipitation (mm)'
        chart.x_axis.title = 'Return Period (Years)'
        chart.style = 2
        n_blocks = len(gumbel_params)
        data = Reference(ws, min_col=2, max_col=1 + n_blocks,
                          min_row=curve_row0, max_row=curve_row0 + len(return_periods_curve))
        cats = Reference(ws, min_col=1, max_col=1,
                          min_row=curve_row0 + 1, max_row=curve_row0 + len(return_periods_curve))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for s in chart.series:
            s.smooth = True
        ws.add_chart(chart, f'{get_column_letter(2 + n_blocks + 1)}{curve_row0}')
    except Exception as e:
        print(f"    [template-excel] return-period curve chart failed: {e}")

    ws.column_dimensions['A'].width = 12


# =========================================================================
# Entry point
# =========================================================================
def write_template_style_excel(out_path, variables, _cfg, scenarios, future_intervals,
                                ref_cache, corrected_grids, results, precip_thresholds, return_periods=None):
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _sheet_daily_spatial_averages(wb, variables, _cfg, scenarios, future_intervals,
                                   ref_cache, corrected_grids)
    _sheet_helper(wb, scenarios, future_intervals, ref_cache, _cfg, corrected_grids)
    _sheet_precipitation_stats(wb, results, scenarios, future_intervals, precip_thresholds, return_periods)
    _sheet_max_precip(wb, scenarios, future_intervals, ref_cache, corrected_grids)
    _sheet_return_period(wb, scenarios, future_intervals, ref_cache, corrected_grids)
    _sheet_temperature_stats(wb, results, scenarios, future_intervals)

    wb.save(out_path)
    return out_path
